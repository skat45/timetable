import openpyxl
from flask import Flask, render_template
import requests
from datetime import date, time

weather_last_update = None
weather_temperature = None
weather_wind = None
delata_update_time = 60 * 30


excel_settings_file = openpyxl.load_workbook('Настройки.xlsx')
excel_file_const_name = excel_settings_file['Параметры'].cell(row=1, column=2).value
excel_file_not_const_name = excel_settings_file['Параметры'].cell(row=2, column=2).value

excel_file_const = openpyxl.load_workbook(excel_file_const_name)
excel_file_not_const = openpyxl.load_workbook(excel_file_not_const_name)
sheets = excel_file_const.sheetnames
sheets_not_const = excel_file_not_const.sheetnames

birthdays_sheet = excel_settings_file['ДР']

openweathermap_key = excel_settings_file['Параметры'].cell(row=4, column=2).value
city_id = excel_settings_file['Параметры'].cell(row=5, column=2).value


def weather_update():
    global delata_update_time, weather_last_update, weather_wind, weather_temperature, city_id, openweathermap_key
    now = [time.hour, time.minute]
    if weather_last_update is None or now[0] * 60 + now[1] - delata_update_time >= weather_last_update[0] * 60 + \
            weather_last_update[1]:
        response = requests.get(
            f'http://api.openweathermap.org/data/2.5/weather?id={city_id}&appid={openweathermap_key}&units=metric').json()
    
    try:
        weather_temperature = response['main']['temp']
        weather_wind = response['wind']['speed']
    except:
        if weather_temperature != None:
            pass
        else:
            weather_temperature = 'error'
            weather_wind = 'error'
    # weather_temperature = '0,34'
    # weather_wind = '12'


weather_update()

app = Flask(__name__, static_folder="templates/static")

Monday = {}  # Пн
Tuesday = {}  # Вт
Wednesday = {}  # Ср
Thursday = {}  # Чт
Friday = {}  # Пт
Saturday = {}  # Сб
Sunday = {}  # Вс

Monday_nc = {}  # Пн
Tuesday_nc = {}  # Вт
Wednesday_nc = {}  # Ср
Thursday_nc = {}  # Чт
Friday_nc = {}  # Пт
Saturday_nc = {}  # Сб
Sunday_nc = {}  # Вс


@app.route('/', methods=['POST', 'GET'])
def index():
    return render_template('index.html')

@app.route('/get_bd_today', methods=['POST', 'GET'])
def birthday_today():
    global birthdays_sheet, excel_settings_file
    excel_settings_file = openpyxl.load_workbook('Настройки.xlsx')
    birthdays_sheet = excel_settings_file['ДР']
    today_day = date.today().day
    today_month = date.today().month
    all_bithdays = []
    max_num = birthdays_sheet.max_row
    for this_row in range(2, max_num + 1):
        try:
            if today_day == int(birthdays_sheet.cell(row=this_row, column=2).value) and today_month == int(
                    birthdays_sheet.cell(row=this_row, column=3).value):
                all_bithdays.append({'name': birthdays_sheet.cell(row=this_row, column=1).value,
                                     'description': birthdays_sheet.cell(row=this_row, column=4).value,
                                     'photo': birthdays_sheet.cell(row=this_row, column=5).value})
        except:
            pass
    return {'data': all_bithdays}


@app.route('/get_weather', methods=['POST', 'GET'])
def get_weather():
    weather_update()
    global weather_temperature, weather_wind
    return {'temperature': weather_temperature, 'wind': weather_wind}


@app.route('/get_const', methods=['POST', 'GET'])
def get_const():
    scan_const()
    data = {1: Monday, 2: Tuesday, 3: Wednesday, 4: Thursday, 5: Friday, 6: Saturday, 0: Sunday}
    return {'data': data}


@app.route('/get_not_const', methods=['POST', 'GET'])
def get_not_const():
    scan_not_const()
    data = {1: Monday_nc, 2: Tuesday_nc, 3: Wednesday_nc, 4: Thursday_nc, 5: Friday_nc, 6: Saturday_nc, 0: Sunday_nc}
    return {'data': data}


def write_this_day(school_class, start_pos, end_pos, day, max_col):
    classes = []
    lessons = {}
    for i in range(2, max_col, 2):
        classes.append(excel_file_const[school_class].cell(row=start_pos, column=i).value)

    for i in range(0, len(classes)):
        lessons_list = []
        for j in range(start_pos + 1, end_pos + 1):
            this_value = excel_file_const[school_class].cell(row=j, column=1).value
            if str(this_value) in '12345678910111213141516171820':
                lessons_list.append([excel_file_const[school_class].cell(row=j, column=(1 + i * 2 + 1)).value,
                                     excel_file_const[school_class].cell(row=j, column=(1 + i * 2 + 2)).value])
            elif str(this_value) in 'К к Классный час классный час':
                lessons_list.append(
                    ['ClassHour', excel_file_const[school_class].cell(row=j, column=(1 + i * 2 + 2)).value])
        lessons[classes[i]] = lessons_list

    if day in 'ПонедельникпнПнПНпонедельник':
        Monday[str(school_class)] = lessons

    elif day in 'ВторниквтВтВТвторник':
        Tuesday[str(school_class)] = lessons

    elif day in 'СредасрСРСрсреда':
        Wednesday[str(school_class)] = lessons

    elif day in 'ЧетвергЧтЧТчтчетверг':
        Thursday[str(school_class)] = lessons

    elif day in 'ПятницаптПтПТпятница':
        Friday[str(school_class)] = lessons

    elif day in 'СубботасбСбСБсуббота':
        Saturday[str(school_class)] = lessons

    elif day in 'ВоскресеньевсВсВСвоскресенье':
        Sunday[str(school_class)] = lessons


def write_this_day_nc(school_class, start_pos, end_pos, day, max_col):
    classes = []
    lessons = {}
    for i in range(2, max_col, 2):
        classes.append(excel_file_not_const[school_class].cell(row=start_pos, column=i).value)

    for i in range(0, len(classes)):
        lessons_list = []
        for j in range(start_pos + 1, end_pos + 1):
            this_value = excel_file_not_const[school_class].cell(row=j, column=1).value
            if str(this_value) in '12345678910111213141516171820':
                lessons_list.append([excel_file_not_const[school_class].cell(row=j, column=(1 + i * 2 + 1)).value,
                                     excel_file_not_const[school_class].cell(row=j, column=(1 + i * 2 + 2)).value])
            elif str(this_value) in 'К к Классный час классный час':
                lessons_list.append(
                    ['ClassHour', excel_file_not_const[school_class].cell(row=j, column=(1 + i * 2 + 2)).value])
        lessons[classes[i]] = lessons_list

    if day in 'ПонедельникпнПнПНпонедельник':
        Monday_nc[str(school_class)] = lessons

    elif day in 'ВторниквтВтВТвторник':
        Tuesday_nc[str(school_class)] = lessons

    elif day in 'СредасрСРСрсреда':
        Wednesday_nc[str(school_class)] = lessons

    elif day in 'ЧетвергЧтЧТчтчетверг':
        Thursday_nc[str(school_class)] = lessons

    elif day in 'ПятницаптПтПТпятница':
        Friday_nc[str(school_class)] = lessons

    elif day in 'СубботасбСбСБсуббота':
        Saturday_nc[str(school_class)] = lessons

    elif day in 'ВоскресеньевсВсВСвоскресенье':
        Sunday_nc[str(school_class)] = lessons


def scan_const():
    global excel_settings_file, excel_file_const_name, excel_file_const, sheets
    excel_settings_file = openpyxl.load_workbook('Настройки.xlsx')
    excel_file_const_name = excel_settings_file['Параметры'].cell(row=1, column=2).value

    excel_file_const = openpyxl.load_workbook(excel_file_const_name)
    sheets = excel_file_const.sheetnames

    for school_class in sheets:
        rows_num = int(excel_file_const[school_class].max_row) + 1
        columns_num = int(excel_file_const[school_class].max_column)

        none_before = True
        start_position = 0
        this_day = ''

        for this_row in range(1, rows_num + 1):
            cell_value = excel_file_const[school_class].cell(row=this_row, column=1).value
            if none_before and cell_value != None:
                none_before = False
                start_position = this_row
                this_day = cell_value

            if cell_value == None:
                none_before = True
                end_position = int(this_row) - 1
                write_this_day(school_class, start_position, end_position, this_day, columns_num)


def scan_not_const():
    global excel_settings_file, excel_file_not_const_name, excel_file_not_const, sheets_not_const
    excel_settings_file = openpyxl.load_workbook('Настройки.xlsx')
    excel_file_not_const_name = excel_settings_file['Параметры'].cell(row=2, column=2).value

    excel_file_not_const = openpyxl.load_workbook(excel_file_not_const_name)
    sheets_not_const = excel_file_not_const.sheetnames

    for school_class in sheets:
        rows_num = int(excel_file_not_const[school_class].max_row) + 1
        columns_num = int(excel_file_not_const[school_class].max_column)

        none_before = True
        start_position = 0
        this_day = ''

        for this_row in range(1, rows_num + 1):
            cell_value = excel_file_not_const[school_class].cell(row=this_row, column=1).value
            if none_before and cell_value != None:
                none_before = False
                start_position = this_row
                this_day = cell_value

            if cell_value == None:
                none_before = True
                end_position = int(this_row) - 1
                write_this_day_nc(school_class, start_position, end_position, this_day, columns_num)


if __name__ == "__main__":
    app.run(debug=True)
